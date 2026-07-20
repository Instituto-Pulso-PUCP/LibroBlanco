#!/usr/bin/env python3
"""Export a pipeline CSV/DataFrame to a readable XLSX whose column headers are
colored by the data source each column comes from.

Sources recognized (see ``SOURCE_DEFS``):

- ``proyecto``   : columnas del Excel original de VRI (hoja PROYECTOS).
- ``vri``        : resultados declarados por VRI (hoja PROY_RESULTADOS).
- ``matching``   : columnas calculadas por el pipeline (enlace a publicaciones).
- ``openalex``   : enriquecimiento OpenAlex (columnas ``openalex_*``).
- ``resumenes``  : resumenes/palabras clave integrados desde "Obtencion de
                   resumenes" (Scopus/PubMed/OpenAlex/Crossref).
- ``otro``       : cualquier columna no clasificada.

Usage (standalone)::

    python export_xlsx.py salidas/06_project_results_ground_truth.csv \
        salidas/06_project_results_ground_truth.xlsx
"""

from __future__ import annotations

import argparse
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# Each source: header fill color, font color, and a human-readable label.
SOURCE_DEFS = {
    'proyecto':  {'fill': '1F4E79', 'font': 'FFFFFF', 'label': 'Proyecto (Excel VRI - hoja PROYECTOS)'},
    'vri':       {'fill': 'C55A11', 'font': 'FFFFFF', 'label': 'Resultado declarado por VRI (PROY_RESULTADOS)'},
    'matching':  {'fill': '7F7F7F', 'font': 'FFFFFF', 'label': 'Enlace calculado por el pipeline'},
    'openalex':  {'fill': '548235', 'font': 'FFFFFF', 'label': 'Enriquecimiento OpenAlex'},
    'resumenes': {'fill': '7030A0', 'font': 'FFFFFF', 'label': 'Resumenes/palabras clave (Scopus/PubMed/OpenAlex/Crossref)'},
    'otro':      {'fill': '404040', 'font': 'FFFFFF', 'label': 'Otra fuente'},
}

# Explicit column -> source assignments for the known pipeline schema.
_PROYECTO_COLS = {
    'project_id', 'cod_aeri', 'codigo_campus', 'project_year', 'project_title',
    'coordinator_original',
}
_VRI_COLS = {
    'cod_prod', 'result_type', 'result_category', 'result_title', 'result_other',
    'result_year', 'equivalencia', 'doi_raw', 'doi_norm', 'issn_raw', 'isbn_raw',
    'result_status', 'result_evidence', 'scopus_eid', 'journal_raw', 'quartile_raw',
    'citation_raw', 'citation_fw_raw', 'citation_fa_raw', 'citation_policy_raw',
    'call_name', 'result_idperson', 'result_coordinator', 'result_is_publication_like',
}
_MATCHING_COLS = {'publication_id', 'match_method'}


def classify_column(column: str) -> str:
    """Return the source key for a column name."""
    col = str(column).strip()
    if col.startswith('openalex_'):
        return 'openalex'
    if col.startswith('resumen') or col.startswith('palabras_clave') or col.startswith('fuente_'):
        return 'resumenes'
    if col in _PROYECTO_COLS:
        return 'proyecto'
    if col in _VRI_COLS:
        return 'vri'
    if col in _MATCHING_COLS:
        return 'matching'
    return 'otro'


def _cell_value(value):
    """Coerce a pandas value into something openpyxl can write."""
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(value, (bool, int, float, str)):
        return value
    return str(value)


def write_colored_xlsx(data, output_path, title=None, max_col_width=60):
    """Write ``data`` (a DataFrame or CSV path) to ``output_path`` with headers
    colored by source, a frozen/auto-filtered header row, and a legend sheet.
    """
    if isinstance(data, (str, Path)):
        df = pd.read_csv(data, dtype=str, keep_default_na=False)
    else:
        df = data.copy()

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = (title or 'Datos')[:31]

    columns = list(df.columns)
    sources = [classify_column(c) for c in columns]

    header_font = Font(bold=True)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Header row
    for col_idx, (col_name, src) in enumerate(zip(columns, sources), start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        style = SOURCE_DEFS[src]
        cell.fill = PatternFill(start_color=style['fill'], end_color=style['fill'], fill_type='solid')
        cell.font = Font(bold=True, color=style['font'])
        cell.alignment = center

    # Data rows
    for row_idx, record in enumerate(df.itertuples(index=False, name=None), start=2):
        for col_idx, value in enumerate(record, start=1):
            ws.cell(row=row_idx, column=col_idx, value=_cell_value(value))

    # Freeze header, enable autofilter
    ws.freeze_panes = 'A2'
    last_col = get_column_letter(len(columns)) if columns else 'A'
    ws.auto_filter.ref = f'A1:{last_col}{max(1, len(df) + 1)}'

    # Column widths based on header + a sample of the data
    sample = df.head(200)
    for col_idx, col_name in enumerate(columns, start=1):
        header_len = len(str(col_name))
        try:
            data_len = int(sample[col_name].astype(str).map(len).max())
        except (ValueError, TypeError):
            data_len = 0
        width = min(max_col_width, max(10, header_len + 2, data_len + 2))
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Legend sheet
    legend = wb.create_sheet('Leyenda')
    legend['A1'] = 'Color'
    legend['B1'] = 'Fuente'
    legend['C1'] = 'Columnas'
    for cell in (legend['A1'], legend['B1'], legend['C1']):
        cell.font = header_font
    used_sources = []
    for src in SOURCE_DEFS:
        if src in sources:
            used_sources.append(src)
    for row_idx, src in enumerate(used_sources, start=2):
        style = SOURCE_DEFS[src]
        swatch = legend.cell(row=row_idx, column=1, value=' ')
        swatch.fill = PatternFill(start_color=style['fill'], end_color=style['fill'], fill_type='solid')
        legend.cell(row=row_idx, column=2, value=style['label'])
        cols_for_src = [c for c, s in zip(columns, sources) if s == src]
        legend.cell(row=row_idx, column=3, value=', '.join(cols_for_src))
    legend.column_dimensions['A'].width = 8
    legend.column_dimensions['B'].width = 55
    legend.column_dimensions['C'].width = 90

    wb.save(output_path)
    return output_path


def main(argv=None):
    parser = argparse.ArgumentParser(
        description='Exporta un CSV del pipeline a XLSX con encabezados coloreados por fuente.')
    parser.add_argument('input_csv', type=Path, help='CSV de entrada.')
    parser.add_argument('output_xlsx', type=Path, nargs='?',
                        help='XLSX de salida (por defecto, mismo nombre con extension .xlsx).')
    parser.add_argument('--title', default=None, help='Titulo de la hoja de datos.')
    args = parser.parse_args(argv)

    if not args.input_csv.exists():
        raise FileNotFoundError(f'No existe el archivo: {args.input_csv}')
    output = args.output_xlsx or args.input_csv.with_suffix('.xlsx')
    path = write_colored_xlsx(args.input_csv, output, title=args.title)
    print(f'Escrito: {path.resolve()}')


if __name__ == '__main__':
    main()
