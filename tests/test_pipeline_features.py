"""Tests for the stop/resume caching, source-colored XLSX export, and the
resumenes merge added to the pipeline."""

import sys
from pathlib import Path

import pandas as pd
import pytest

ROOT = Path(__file__).resolve().parents[1] / "scripts"
sys.path.insert(0, str(ROOT))

import openalex_helpers as oh
from export_xlsx import classify_column, write_colored_xlsx
import merge_resumenes as mr


# --------------------------------------------------------------------------
# Caching / resume
# --------------------------------------------------------------------------

def test_enrichment_cache_key_prefers_doi_and_normalizes():
    assert oh.enrichment_cache_key(doi="10.1/ABC", title="x") == "doi:10.1/abc"
    assert oh.enrichment_cache_key(doi="", title="Hello   World") == "title:hello world"
    assert oh.enrichment_cache_key(doi="-", title="") == ""


def test_openalex_cache_roundtrip_and_persistence(tmp_path):
    path = tmp_path / "cache.jsonl"
    cache = oh.OpenAlexCache(path)
    assert len(cache) == 0
    cache.put("doi:10.1/x", {"openalex_work_id": "W1", "openalex_enrichment_error": ""})
    assert "doi:10.1/x" in cache
    # A fresh cache reads back the persisted record (simulates resume).
    reloaded = oh.OpenAlexCache(path)
    assert reloaded.get("doi:10.1/x")["openalex_work_id"] == "W1"


def test_fetch_cached_serves_cache_and_skips_transient(monkeypatch, tmp_path):
    cache = oh.OpenAlexCache(tmp_path / "cache.jsonl")

    # First: a transient (429) failure must NOT be cached so it can retry.
    monkeypatch.setattr(oh, "fetch_openalex_enrichment",
                        lambda doi=None, title=None, timeout=10: {"openalex_enrichment_error": "HTTP Error 429"})
    result, was_cached = oh.fetch_openalex_enrichment_cached(doi="10.1/x", cache=cache)
    assert was_cached is False
    assert "doi:10.1/x" not in cache

    # Then: a successful result IS cached and served on the next call.
    monkeypatch.setattr(oh, "fetch_openalex_enrichment",
                        lambda doi=None, title=None, timeout=10: {"openalex_work_id": "W9", "openalex_enrichment_error": ""})
    result, was_cached = oh.fetch_openalex_enrichment_cached(doi="10.1/x", cache=cache)
    assert was_cached is False and result["openalex_work_id"] == "W9"
    result2, was_cached2 = oh.fetch_openalex_enrichment_cached(doi="10.1/x", cache=cache)
    assert was_cached2 is True and result2["openalex_work_id"] == "W9"


# --------------------------------------------------------------------------
# XLSX export
# --------------------------------------------------------------------------

def test_classify_column_maps_known_sources():
    assert classify_column("project_id") == "proyecto"
    assert classify_column("result_title") == "vri"
    assert classify_column("publication_id") == "matching"
    assert classify_column("openalex_cited_by_count") == "openalex"
    assert classify_column("resumen") == "resumenes"
    assert classify_column("palabras_clave") == "resumenes"
    assert classify_column("totally_unknown") == "otro"


def test_write_colored_xlsx_colors_headers_by_source(tmp_path):
    from openpyxl import load_workbook

    df = pd.DataFrame({
        "project_id": [1],
        "result_title": ["A paper"],
        "openalex_is_oa": [True],
        "resumen": ["some abstract"],
    })
    out = tmp_path / "out.xlsx"
    write_colored_xlsx(df, out, title="T")
    wb = load_workbook(out)
    assert "Leyenda" in wb.sheetnames
    ws = wb[wb.sheetnames[0]]
    colors = {ws.cell(1, c).value: ws.cell(1, c).fill.start_color.rgb for c in range(1, 5)}
    # Each source group has a distinct header fill.
    assert len({colors["project_id"], colors["result_title"],
                colors["openalex_is_oa"], colors["resumen"]}) == 4


# --------------------------------------------------------------------------
# Resumenes merge
# --------------------------------------------------------------------------

def test_norm_doi_variants():
    assert mr.norm_doi("https://doi.org/10.1/ABC") == "10.1/abc"
    assert mr.norm_doi("doi: 10.2/x ") == "10.2/x"
    assert mr.norm_doi("-") == ""
    assert mr.norm_doi(None) == ""


def test_merge_into_fills_abstract_columns(tmp_path):
    main = pd.DataFrame({
        "project_id": [1, 2],
        "doi_norm": ["10.1/a", "10.2/b"],
        "result_title": ["t1", "t2"],
    })
    main_csv = tmp_path / "06.csv"
    main.to_csv(main_csv, index=False)

    resumenes = pd.DataFrame({
        "doi": ["https://doi.org/10.1/A"],
        "resumen": ["abstract one"],
        "fuente_resumen": ["OpenAlex"],
        "palabras_clave": ["k1 | k2"],
    })
    # Use the real loader path via a temp csv to exercise consolidation.
    res_csv = tmp_path / "doi-resultados.csv"
    resumenes.to_csv(res_csv, index=False)
    resumenes_df = mr.load_resumenes(res_csv)

    out_csv = tmp_path / "06_con_resumenes.csv"
    stats = mr.merge_into(main_csv, resumenes_df, out_csv, make_xlsx=False)
    merged = pd.read_csv(out_csv, dtype=str, keep_default_na=False)

    assert stats["rows"] == 2
    assert stats["rows_with_abstract_data"] == 1
    row1 = merged[merged["project_id"] == "1"].iloc[0]
    assert row1["resumen"] == "abstract one"
    assert row1["palabras_clave"] == "k1 | k2"
    # Non-matching row stays blank rather than NaN.
    row2 = merged[merged["project_id"] == "2"].iloc[0]
    assert row2["resumen"] == ""
